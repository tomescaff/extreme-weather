"""CONAF's data access functions.

This module contains functions to access CONAF's data stored in CR2's 
data servers. It contains the following main functions:

    * burned_area: Access burned area data over Chilean territory from 
        CONAF's historical records (1964-2022) plus 2023 data from 
        personal communication.
        
    * nfires: Access number of fires data over Chilean territory from
        CONAF's historical records (1964-2022) plus 2023 data from
        personal communication.
        
    * burned_area_valparaiso_araucania: Access burned area data over 
        Chilean territory from Valparaiso to La Araucania regions from
        CONAF's historical records (1977-2022) plus 2023 data from
        personal communication.
        
    * stats_ohiggins_araucania: Access burned area and number of fires 
        over Chilean territory from 1976 to 2022 from unknown source.
        
    * burned_area_per_month_per_region_from_ohiggins_to_araucania: 
        Access monthly burned area data per region from O'Higgins to La 
        Araucania from CONAF's historical records (1985-2023).
        
    * nfires_per_month_per_region_from_ohiggins_to_araucania: Access
        monthly number of fires data per region from O'Higgins to La 
        Araucania from CONAF's historical records (1985-2023). 
"""

import pandas as pd
import numpy as np
import xarray as xr
import openpyxl


def burned_area():
    """Access burned area data over Chilean territory from CONAF's
    historical records (1964-2022) plus 2023 data from personal comm.

    Returns
    -------
    xr.DataArray
        burned area data over Chilean territory from 1963/64 to 2022/23 
        fire seasons (July to June).
    """
    basedir = "/home/tcarrasco/result/data/CONAF/"
    filename = "resumen_nacional_ocurrencia_dano_1964_2022.xlsx"
    filepath = basedir + filename
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook['Historico']
    years = np.arange(1964, 2022+1)
    n_years = years.size
    data = np.zeros((n_years+1,))
    for i in range(n_years):
        cell_value = sheet.cell(row=11+i, column=6).value
        if cell_value is not None:
            try:
                data[i] = float(cell_value) if isinstance(
                    cell_value, (int, float)) else np.nan
            except (ValueError, TypeError):
                data[i] = np.nan
        else:
            data[i] = np.nan
    data[n_years] = 432000
    dr = pd.date_range('1964-01-01', '2023-01-01', freq='1YS')
    da = xr.DataArray(data, coords=[dr], dims=['time'])
    return da


def nfires():
    """Access number of fires data over Chilean territory from CONAF's
    historical records (1964-2022) plus 2023 data from personal comm.

    Returns
    -------
    xr.DataArray
        number of fires data over Chilean territory from 1963/64 to 
        2022/23 fire seasons (July to June).
    """
    basedir = "/home/tcarrasco/result/data/CONAF/"
    filename = "resumen_nacional_ocurrencia_dano_1964_2022.xlsx"
    filepath = basedir + filename
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook['Historico']
    years = np.arange(1964, 2022+1)
    n_years = years.size
    data = np.zeros((n_years+1,))
    for i in range(n_years):
        cell_value = sheet.cell(row=11+i, column=5).value
        if cell_value is not None:
            try:
                data[i] = float(cell_value) if isinstance(
                    cell_value, (int, float)) else np.nan
            except (ValueError, TypeError):
                data[i] = np.nan
        else:
            data[i] = np.nan
    data[n_years] = 4500
    dr = pd.date_range('1964-01-01', '2023-01-01', freq='1YS')
    da = xr.DataArray(data, coords=[dr], dims=['time'])
    return da


def burned_area_valparaiso_araucania():
    """Access burned area data over Chilean territory from Valparaiso to 
    La Araucania regions from CONAF's historical records (1977-2022) 
    plus 2023 data from personal communication.

    Returns
    -------
    xr.DataArray
        burned area data from 1976/77 to 2022/23 fire seasons (July to 
        June) over Chilean territory from Valparaiso to La Araucania.
    """
    basedir = "/home/tcarrasco/result/data/CONAF/"
    filename = "resumen_regional_ocurrencia_dano_1977_2022.xlsx"
    filepath = basedir + filename
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook['Daño']
    years = np.arange(1977, 2022+1)
    n_years = years.size
    m = 8
    data = np.zeros((n_years+1, m))
    for j in range(m):
        for i in range(n_years):
            cell_value = sheet.cell(row=10+i, column=6+j).value
            if cell_value is not None:
                try:
                    data[i, j] = float(cell_value) if isinstance(
                        cell_value, (int, float)) else np.nan
                except (ValueError, TypeError):
                    data[i, j] = np.nan
            else:
                data[i, j] = np.nan
    data_sum = np.nansum(data, axis=1)
    data_sum[n_years] = 432000
    dr = pd.date_range('1977-01-01', '2023-01-01', freq='1YS')
    da = xr.DataArray(data_sum, coords=[dr], dims=['time'])
    return da


def stats_ohiggins_araucania():
    """Access burned area and number of fires over Chilean territory 
    from 1976 to 2022 from unknown source.

    Returns
    -------
    xr.Dataset
        burned area and number of fires data from 1975/76 to 2021/22
        fire seasons (Jluy to June) over Chilean territory from O'Higgins
        to La Araucania.
    """
    basedir = "/home/tcarrasco/result/data/CONAF/"
    filename = "datos_incendios_ROH_RLA.csv"
    filepath = basedir + filename
    df = pd.read_csv(filepath, index_col=None, parse_dates={'time': ['year']})
    df = df.set_index('time')
    ds = df.to_xarray()
    return ds


def region_column(region, year):
    """Auxiliary function to get the column index for a given region and
    year in the CONAF's monthly data spreadsheets.

    Parameters
    ----------
    region: str
        region number in roman numerals
    year: int
        year for which the column index is needed.
    
    Returns
    -------
    int
        column index for the given region and year.
    """
    coldict_2 = {'VI': 9, 'VII': 10, 'XVI': 11, 'VIII': 12, 'IX': 13}
    coldict_1 = {'VI': 9, 'VII': 10, 'XVI': np.nan, 'VIII': 11, 'IX': 12}
    coldict_0 = {'VI': 6, 'VII': 7, 'XVI': np.nan, 'VIII': 8, 'IX': 9}
    if year >= 2019:
        return coldict_2[region]
    if year <= 2016:
        return coldict_0[region]
    return coldict_1[region]


def get_workbook_data(workbook, init_row):
    """Auxiliary function to get the burned area or number of fires data
    from CONAF's monthly data spreadsheets.
    
    Parameters
    ----------
    workbook: openpyxl.workbook.workbook.Workbook
        workbook object representing the CONAF's spreadsheet.
    init_row: int
        initial row in the spreadsheet where the data starts.
        
    Returns
    -------
    xr.DataArray
        fire activity per month per region from O'Higgins to La Araucania 
        from 1984/85 to 2022/23 fire seasons (July to June).
    """
    regions = ['VI', 'VII', 'XVI', 'VIII', 'IX']
    time = pd.date_range('1984-07', '2023-06', freq='1MS')
    mat = np.zeros((len(regions), time.size))
    da = xr.DataArray(mat, coords=[regions, time], dims=['region', 'time'])
    k = 0
    for year in range(1985, 2024):
        # get the sheet for the year
        sheet = workbook[str(year)]
        for i, region in enumerate(regions):
            # get the column for the region and year
            column = region_column(region, year)
            if np.isnan(column):
                continue
            for j in range(12):
                cell_value = sheet.cell(row=init_row + j, column=column).value
                if cell_value is not None:
                    bool_value = isinstance(cell_value, (int, float))
                    da[i, k+j] = float(cell_value) if bool_value else np.nan
                else:
                    da[i, k+j] = np.nan
        k = k + 12
    return da


def burned_area_per_month_per_region_from_ohiggins_to_araucania():
    """Access monthly burned area data per region from O'Higgins to La 
    Araucania from CONAF's historical records (1985-2023).

    Returns
    -------
    xr.DataArray
        burned area per month per region from O'Higgins to La Araucania
        from 1984/85 to 2022/23 fire seasons (July to June).
    """
    basedir = "/home/tcarrasco/result/data/CONAF/"
    filename = "CONAF_area_quemada_por_mes_1985_2023.xlsx"
    workbook = openpyxl.load_workbook(basedir + filename)

    ba = get_workbook_data(workbook, init_row=10)
    return ba


def nfires_per_month_per_region_from_ohiggins_to_araucania():
    """Access monthly number of fires data per region from O'Higgins to 
    La Araucania from CONAF's historical records (1985-2023).

    Returns
    -------
    xr.DataArray
        number of fires per month per region from O'Higgins to La 
        Araucania from 1984/85 to 2022/23 fire seasons (July to June).
    """
    basedir = "/home/tcarrasco/result/data/CONAF/"
    filename = "CONAF_ocurrencia_por_mes_1985_2023.xlsx"
    workbook = openpyxl.load_workbook(basedir + filename)

    nf = get_workbook_data(workbook, init_row=11)
    return nf
